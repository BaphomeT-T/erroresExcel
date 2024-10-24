import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# Función para validar los correos electrónicos que cumplan con el formato NOMBRE.APELLIDO@universidad.edu.ec
def validar_correo(row):
    correo_regex = r'^[A-Z]+\.[A-Z]+@universidad\.edu\.ec$'  # Formato requerido (en mayúsculas)
    correo = row.get('correo_est')
    
    nombre_est = str(row.get('nombre_est', '')).split()[0]
    apellidos = str(row.get('apellido_est', '')).split()
    
    if len(apellidos) < 1:
        return False
    
    primer_apellido = apellidos[0]
    
    if isinstance(correo, str) and re.match(correo_regex, correo, re.IGNORECASE):
        correo_nombre, correo_apellido = correo.split('@')[0].split('.')
        return (correo_nombre.upper() == nombre_est.upper() and correo_apellido.upper() == primer_apellido.upper())
    
    return False

# Función para validar si la fecha tiene el formato "YYYY-MM-DD HH:MM:SS" como se extra de excel
def validar_fecha_nacimiento(fecha):
    formato = "%Y-%m-%d %H:%M:%S"
    try:
        datetime.strptime(str(fecha), formato)
        return True
    except ValueError:
        return False

# Función para calcular la edad a partir de la fecha de nacimiento usando como fecha actual el 01-12-2020
def calcular_edad(fecha_nacimiento):
    try:
        fecha_nacimiento = datetime.strptime(str(fecha_nacimiento).split()[0], "%Y-%m-%d")
        fecha_referencia = datetime(2020, 12, 1)
        return fecha_referencia.year - fecha_nacimiento.year - ((fecha_referencia.month, fecha_referencia.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    except (ValueError, IndexError):
        return None

# Función para validar que el número de teléfono tenga  7 dígitos
def validar_telefono(telefono):
    if pd.isna(telefono):
        return False
    
    if isinstance(telefono, float):
        telefono = str(int(telefono))
    elif isinstance(telefono, str):
        telefono = telefono.strip()
    
    if len(telefono) == 7 and telefono.isdigit():
        return True
    return False

# Función para validar el código de estudiante, debe tener 11 digitos, empezar por 2019 y estar de forma consecutiva
def validar_codigo_est(codigos_est):
    errores = []
    for i, codigo in enumerate(codigos_est):
        if not isinstance(codigo, str) or len(codigo) != 11 or not codigo.startswith('2019'):
            errores.append((i, codigo))
        else:
            consecutivo_actual = 20190000001 + i
            if int(codigo) != consecutivo_actual:
                errores.append((i, codigo))
    return errores

# Función para validar haya un solo nombre en mayúsculas y con caracteres del alfabeto español
def validar_nombre(nombre):
    if pd.isna(nombre):
        return False
    
    nombre = str(nombre).strip()
    
    nombre_regex = r'^[A-ZÁÉÍÓÚÜÑ]+$'
    
    if len(nombre.split()) == 1 and re.match(nombre_regex, nombre):
        return True
    return False

# Función para validar que haya 2 apellidos en mayúsculas y con caracteres del alfabeto español
def validar_apellidos(apellidos):
    if pd.isna(apellidos):
        return False
    
    apellidos = str(apellidos).strip()
    
    apellido_regex = r'^[A-ZÁÉÍÓÚÜÑ]+$'
    
    apellidos_lista = apellidos.split()
    
    if len(apellidos_lista) == 2 and all(re.match(apellido_regex, apellido) for apellido in apellidos_lista):
        return True
    return False

# Función para validar la cédula ecuatoriana 
def validar_cedula(cedula):
    if pd.isna(cedula):
        return False
    
    if isinstance(cedula, float):
        cedula = str(int(cedula))
    
    cedula = cedula.zfill(10)
    
    if len(cedula) != 10 or not cedula.isdigit():
        return False

    provincia = int(cedula[:2])
    if provincia < 0 or provincia > 24:
        return False

    tercer_digito = int(cedula[2])
    if tercer_digito >= 6:
        return False

    coeficientes = [2, 1, 2, 1, 2, 1, 2, 1, 2]
    suma = 0

    for i in range(9):
        producto = int(cedula[i]) * coeficientes[i]
        if producto >= 10:
            producto -= 9
        suma += producto

    digito_verificador = (10 - (suma % 10)) % 10

    if digito_verificador != int(cedula[9]):
        return False

    return True

# Función para validar que la direccion cumpla con el formato Calle1 número Calle2, calle1 y 2 pueden ser cualquier tipo de strings
def validar_direccion(direccion):
    if pd.isna(direccion):
        return False
    direccion = str(direccion).strip().upper()
    # La dirección debe seguir el patrón "Calle 1 número Calle 2" con múltiples palabras y separadores
    direccion_regex = r'^[A-Z0-9\s\.,]+\s+\d+\s+[A-Z0-9\s\.,]+$'
    return bool(re.match(direccion_regex, direccion))



# Función principal 
def analizar_calidad_datos(file_path):
    df = pd.read_excel(file_path)
    wb = load_workbook(file_path)
    ws = wb.active

    errores = {col: 0 for col in df.columns}

    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for idx, row in df.iterrows():
        for col in df.columns:
            if pd.isna(row[col]):
                cell = ws.cell(row=idx+2, column=df.columns.get_loc(col)+1)
                cell.fill = fill_red
                errores[col] += 1

        if 'correo_est' in df.columns and not pd.isna(row.get('correo_est')) and not validar_correo(row):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('correo_est')+1)
            cell.fill = fill_red
            errores['correo_est'] += 1

        if 'calificacion' in df.columns and not pd.isna(row.get('calificacion')) and not (0 <= row.get('calificacion', 0) <= 20):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('calificacion')+1)
            cell.fill = fill_red
            errores['calificacion'] += 1

        if 'fecha_nacimiento_est' in df.columns and not pd.isna(row.get('fecha_nacimiento_est')) and not validar_fecha_nacimiento(row.get('fecha_nacimiento_est', '')):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('fecha_nacimiento_est')+1)
            cell.fill = fill_red
            errores['fecha_nacimiento_est'] += 1

        if 'fecha_nacimiento_est' in df.columns and 'edad_est' in df.columns:
            edad_calculada = calcular_edad(row.get('fecha_nacimiento_est'))
            edad_reportada = row.get('edad_est')
            if edad_reportada != edad_calculada:
                cell = ws.cell(row=idx+2, column=df.columns.get_loc('edad_est')+1)
                cell.fill = fill_red
                errores['edad_est'] += 1

        if 'telefono_est' in df.columns and not pd.isna(row.get('telefono_est')) and not validar_telefono(row.get('telefono_est', '')):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('telefono_est')+1)
            cell.fill = fill_red
            errores['telefono_est'] += 1
        
        if 'nombre_est' in df.columns and not pd.isna(row.get('nombre_est')) and not validar_nombre(row.get('nombre_est', '')):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('nombre_est')+1)
            cell.fill = fill_red
            errores['nombre_est'] += 1
        
        if 'apellido_est' in df.columns and not pd.isna(row.get('apellido_est')) and not validar_apellidos(row.get('apellido_est', '')):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('apellido_est')+1)
            cell.fill = fill_red
            errores['apellido_est'] += 1
            
        if 'cedula_est' in df.columns and not pd.isna(row.get('cedula_est')) and not validar_cedula(row.get('cedula_est', '')):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('cedula_est')+1)
            cell.fill = fill_red
            errores['cedula_est'] += 1
            
        if 'direccion_est' in df.columns and not validar_direccion(row.get('direccion_est', '')):
            cell = ws.cell(row=idx+2, column=df.columns.get_loc('direccion_est')+1)
            cell.fill = fill_red
            errores['direccion_est'] += 1


    if 'codigo_est' in df.columns:
        codigos_est = df['codigo_est'].astype(str).tolist()
        errores_codigo_est = validar_codigo_est(codigos_est)
        for i, codigo in errores_codigo_est:
            cell = ws.cell(row=i+2, column=df.columns.get_loc('codigo_est')+1)
            cell.fill = fill_red
            errores['codigo_est'] += 1

    output_path = file_path.replace(".xlsx", "_validado.xlsx")
    wb.save(output_path)

    return errores, output_path

# Ejecutar la función con el archivo de excel
file_path = '2020A_calidad_datos.xlsx'  
errores, archivo_validado = analizar_calidad_datos(file_path)

print("Errores por columna:", errores)
print(f"Archivo validado guardado en: {archivo_validado}")
